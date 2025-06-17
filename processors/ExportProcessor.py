import pandas as pd
import numpy as np
import os
import zipfile
from datetime import datetime
import re
from shapely.geometry import Polygon, Point, MultiPolygon

class ExportProcessor:
    def __init__(self):
        pass

    def export_small_tiles_report(self, small_tiles_df, size_threshold=10, output_prefix="small_tiles"):
        """Export small tiles data to files"""
        if len(small_tiles_df) == 0:
            print("\n📝 No small tiles to export")
            return []
        
        print("\n📝 Exporting small tiles report...")
        export_files = []
        
        # Create a summary DataFrame with useful information
        summary_data = []
        for _, tile in small_tiles_df.iterrows():
            summary_data.append({
                'Apartment': tile['apartment_name'],
                'Room': tile['room_name'],
                'Classification': tile['classification'],
                'Cut Dimension (mm)': tile['cut_side'],
                'Measured Width (mm)': round(tile['measured_width'], 1),
                'Measured Height (mm)': round(tile['measured_height'], 1),
                'Tile Index': tile['tile_index']
            })
        
        # Convert to DataFrame and sort
        summary_df = pd.DataFrame(summary_data)
        summary_df = summary_df.sort_values(['Apartment', 'Room', 'Cut Dimension (mm)'])
        
        # Export to CSV
        csv_file = f"{output_prefix}_under_{size_threshold}mm.csv"
        summary_df.to_csv(csv_file, index=False)
        print(f"✅ Exported small tiles report to {csv_file}")
        export_files.append(csv_file)
        
        # Try to export to Excel
        try:
            excel_file = f"{output_prefix}_under_{size_threshold}mm.xlsx"
            
            with pd.ExcelWriter(excel_file) as writer:
                # Write summary sheet
                summary_df.to_excel(writer, sheet_name='Small Tiles List', index=False)
                
                # Create a pivot table by location
                pivot_df = pd.pivot_table(
                    summary_df, 
                    values='Tile Index', 
                    index=['Apartment', 'Room'], 
                    aggfunc='count'
                ).reset_index()
                pivot_df.rename(columns={'Tile Index': 'Count'}, inplace=True)
                pivot_df.to_excel(writer, sheet_name='By Location', index=False)
                
                # Create bins for the cut dimensions
                bins = [0, 2, 4, 6, 8, 10]
                labels = ['0-2mm', '2-4mm', '4-6mm', '6-8mm', '8-10mm']
                summary_df['Size Range'] = pd.cut(summary_df['Cut Dimension (mm)'], bins=bins, labels=labels)
                
                # Create size distribution pivot
                size_dist = pd.pivot_table(
                    summary_df,
                    values='Tile Index',
                    index=['Size Range'],
                    aggfunc='count'
                ).reset_index()
                size_dist.rename(columns={'Tile Index': 'Count'}, inplace=True)
                size_dist.to_excel(writer, sheet_name='Size Distribution', index=False)
            
            print(f"✅ Exported detailed Excel report to {excel_file}")
            export_files.append(excel_file)
        except Exception as e:
            print(f"⚠️ Could not export to Excel: {e}")
        
        return export_files

    def export_remaining_tiles_with_wastage_analysis(self, tiles_df, small_tiles_df, final_room_df, 
                                                   size_threshold=10, output_prefix="final_tiles_export"):
        """Export remaining tiles with small cuts summary and wastage analysis"""
        print("\n🔄 Preparing export data...")

        # Use the current directory for saving files
        export_path = os.getcwd()
        print(f"📁 Save location: {export_path}")
        
        # Get pattern mode from the tiles classification
        has_pattern = 'cut_x' in tiles_df['classification'].unique() or 'cut_y' in tiles_df['classification'].unique()

        # Process full and irregular tiles
        print("\n🔄 Processing full and irregular tiles...")
        full_df = tiles_df[tiles_df['classification'] == 'full'].copy()
        irregular_df = tiles_df[tiles_df['classification'] == 'irregular'].copy()
        full_and_irregular = pd.concat([full_df, irregular_df])
        
        # Create full tiles summary
        if not full_and_irregular.empty:
            apt_counts = full_and_irregular.groupby('apartment_name').size().reset_index(name='COUNT')
            full_tiles_summary = []
            
            for _, row in apt_counts.iterrows():
                apt_name = row['apartment_name']
                has_irregular = apt_name in irregular_df['apartment_name'].values
                full_tiles_summary.append({
                    'LEVEL': 'Level 0',
                    'APPATMENT NO': apt_name,
                    'TYPE': 'TL-1',
                    'COUNT': row['COUNT'],
                    'LOCATION(room name)': '',
                    'Remarks': "Full + irregular" if has_irregular else ""
                })
            
            full_tiles_export = pd.DataFrame(full_tiles_summary)
        else:
            full_tiles_export = pd.DataFrame(columns=[
                'LEVEL', 'APPATMENT NO', 'TYPE', 'COUNT', 'LOCATION(room name)', 'Remarks'
            ])
        
        # Process cut tiles based on pattern mode
        if has_pattern:
            print("\n🔄 Processing cut X and Y tiles...")
            
            # Process cut_x tiles
            cut_x_df = tiles_df[tiles_df['classification'] == 'cut_x'].copy()
            cut_x_export = pd.DataFrame()
            
            if not cut_x_df.empty:
                # Detailed version with tile sizes
                cut_x_summary = [{
                    'APPARTMENT NUMBER': tile['apartment_name'],
                    'TILE TYPE': 'TL-1',
                    'CUT DIRECTION': 'X_DIRECTION',
                    'TILE SIZE(mm)': f"{round(tile['cut_side'])} x {round(tile['measured_height'])}",
                    'CUT SIDE': round(tile['cut_side']),
                    'LOCATION': tile['room_name'],
                    'COUNT': 1
                } for _, tile in cut_x_df.iterrows()]
                
                cut_x_export = pd.DataFrame(cut_x_summary)
                if not cut_x_export.empty:
                    cut_x_export = cut_x_export.groupby([
                        'APPARTMENT NUMBER', 'TILE TYPE', 'CUT DIRECTION', 
                        'TILE SIZE(mm)', 'CUT SIDE', 'LOCATION'
                    ]).sum().reset_index().sort_values(['APPARTMENT NUMBER', 'CUT SIDE'])
                
                # Simplified version without tile sizes
                cut_x_simple_summary = [{
                    'APPARTMENT NUMBER': tile['apartment_name'],
                    'CUT SIDE (mm)': round(tile['cut_side']),
                    'LOCATION': tile['room_name'],
                    'COUNT': 1
                } for _, tile in cut_x_df.iterrows()]
                
                cut_x_simple = pd.DataFrame(cut_x_simple_summary)
                if not cut_x_simple.empty:
                    cut_x_simple = cut_x_simple.groupby([
                        'APPARTMENT NUMBER', 'CUT SIDE (mm)', 'LOCATION'
                    ]).sum().reset_index().sort_values(['APPARTMENT NUMBER', 'CUT SIDE (mm)'])
            else:
                cut_x_simple = pd.DataFrame(columns=['APPARTMENT NUMBER', 'CUT SIDE (mm)', 'LOCATION', 'COUNT'])
            
            # Process cut_y tiles (similar to cut_x)
            cut_y_df = tiles_df[tiles_df['classification'] == 'cut_y'].copy()
            cut_y_export = pd.DataFrame()
            
            if not cut_y_df.empty:
                # Detailed version with tile sizes
                cut_y_summary = [{
                    'APPARTMENT NUMBER': tile['apartment_name'],
                    'TILE TYPE': 'TL-1',
                    'CUT DIRECTION': 'Y_DIRECTION',
                    'TILE SIZE(mm)': f"{round(tile['measured_width'])} x {round(tile['cut_side'])}",
                    'CUT SIDE': round(tile['cut_side']),
                    'LOCATION': tile['room_name'],
                    'COUNT': 1
                } for _, tile in cut_y_df.iterrows()]
                
                cut_y_export = pd.DataFrame(cut_y_summary)
                if not cut_y_export.empty:
                    cut_y_export = cut_y_export.groupby([
                        'APPARTMENT NUMBER', 'TILE TYPE', 'CUT DIRECTION', 
                        'TILE SIZE(mm)', 'CUT SIDE', 'LOCATION'
                    ]).sum().reset_index().sort_values(['APPARTMENT NUMBER', 'CUT SIDE'])
                
                # Simplified version without tile sizes
                cut_y_simple_summary = [{
                    'APPARTMENT NUMBER': tile['apartment_name'],
                    'CUT SIDE (mm)': round(tile['cut_side']),
                    'LOCATION': tile['room_name'],
                    'COUNT': 1
                } for _, tile in cut_y_df.iterrows()]
                
                cut_y_simple = pd.DataFrame(cut_y_simple_summary)
                if not cut_y_simple.empty:
                    cut_y_simple = cut_y_simple.groupby([
                       'APPARTMENT NUMBER', 'CUT SIDE (mm)', 'LOCATION'
                   ]).sum().reset_index().sort_values(['APPARTMENT NUMBER', 'CUT SIDE (mm)'])
            else:
                cut_y_simple = pd.DataFrame(columns=['APPARTMENT NUMBER', 'CUT SIDE (mm)', 'LOCATION', 'COUNT'])
            
            all_cut_export = pd.DataFrame()  # Empty for pattern mode
            all_cut_simple = pd.DataFrame()  # Empty for pattern mode
            
        else:
            # Process all_cut tiles (no pattern)
            print("\n🔄 Processing all cut tiles (no pattern)...")
            all_cut_df = tiles_df[tiles_df['classification'] == 'all_cut'].copy()
            all_cut_export = pd.DataFrame()
            
            if not all_cut_df.empty:
                all_cut_summary = [{
                    'APPARTMENT NUMBER': tile['apartment_name'],
                    'TILE TYPE': 'TL-1',
                    'CUT DIMENSION': round(tile['cut_side']),
                    'LOCATION': tile['room_name'],
                    'COUNT': 1
                } for _, tile in all_cut_df.iterrows()]
                
                all_cut_export = pd.DataFrame(all_cut_summary)
                if not all_cut_export.empty:
                    all_cut_export = all_cut_export.groupby([
                        'APPARTMENT NUMBER', 'TILE TYPE', 'CUT DIMENSION', 'LOCATION'
                    ]).sum().reset_index().sort_values(['APPARTMENT NUMBER', 'CUT DIMENSION'])
                
                # Simplified version
                all_cut_simple_summary = [{
                    'APPARTMENT NUMBER': tile['apartment_name'],
                    'CUT DIMENSION (mm)': round(tile['cut_side']),
                    'LOCATION': tile['room_name'],
                    'COUNT': 1
                } for _, tile in all_cut_df.iterrows()]
                
                all_cut_simple = pd.DataFrame(all_cut_simple_summary)
                if not all_cut_simple.empty:
                    all_cut_simple = all_cut_simple.groupby([
                        'APPARTMENT NUMBER', 'CUT DIMENSION (mm)', 'LOCATION'
                    ]).sum().reset_index().sort_values(['APPARTMENT NUMBER', 'CUT DIMENSION (mm)'])
            else:
                all_cut_simple = pd.DataFrame(columns=['APPARTMENT NUMBER', 'CUT DIMENSION (mm)', 'LOCATION', 'COUNT'])
            
            cut_x_export = pd.DataFrame()  # Empty for no pattern
            cut_y_export = pd.DataFrame()  # Empty for no pattern
            cut_x_simple = pd.DataFrame()  # Empty for no pattern
            cut_y_simple = pd.DataFrame()  # Empty for no pattern
        
        # Create small cuts summary
        print("\n🔄 Creating small cut tiles summary...")
        small_cuts_export = pd.DataFrame()
        
        if not small_tiles_df.empty:
            if has_pattern:
                # Group by apartment and classification for pattern mode
                small_by_type = small_tiles_df.groupby(['apartment_name', 'classification']).size().unstack(fill_value=0).reset_index()
                
                # Add missing columns if needed
                if 'cut_x' not in small_by_type.columns:
                    small_by_type['cut_x'] = 0
                if 'cut_y' not in small_by_type.columns:
                    small_by_type['cut_y'] = 0
                    
                # Rename columns for clarity
                small_by_type.rename(columns={'cut_x': 'Small X Cuts', 'cut_y': 'Small Y Cuts'}, inplace=True)
                small_by_type['Total Small Cuts'] = small_by_type['Small X Cuts'] + small_by_type['Small Y Cuts']
                
                # Create summary dataframe
                small_cuts_summary = [{
                    'APPARTMENT NUMBER': row['apartment_name'],
                    'SMALL X CUTS COUNT': row['Small X Cuts'],
                    'SMALL Y CUTS COUNT': row['Small Y Cuts'],
                    'TOTAL SMALL CUTS': row['Total Small Cuts'],
                    'SIZE THRESHOLD': f"<{size_threshold}mm",
                    'STATUS': 'REMOVED'
                } for _, row in small_by_type.iterrows()]
                
                small_cuts_export = pd.DataFrame(small_cuts_summary)
                
            else:
                # Simpler summary for no pattern mode
                small_by_apt = small_tiles_df.groupby('apartment_name').size().reset_index(name='Total Small Cuts')
                
                small_cuts_summary = [{
                    'APPARTMENT NUMBER': row['apartment_name'],
                    'TOTAL SMALL CUTS': row['Total Small Cuts'],
                    'SIZE THRESHOLD': f"<{size_threshold}mm",
                    'STATUS': 'REMOVED'
                } for _, row in small_by_apt.iterrows()]
                
                small_cuts_export = pd.DataFrame(small_cuts_summary)
        
        # Calculate area by apartment (include small tiles)
        print("\n🔄 Processing area calculations...")
        
        # Combine regular and small tiles for area calculation
        all_tiles_for_area = pd.concat([tiles_df, small_tiles_df]) if not small_tiles_df.empty else tiles_df
        
        # Calculate area by apartment
        area_by_apt = []
        for apt_name in sorted(all_tiles_for_area['apartment_name'].unique()):
            apt_df = all_tiles_for_area[all_tiles_for_area['apartment_name'] == apt_name]
            
            # Sum up all tile areas for this apartment
            total_area = sum(
                tile['polygon'].area 
                for _, tile in apt_df.iterrows() 
                if isinstance(tile['polygon'], (Polygon, MultiPolygon))
            )
            
            # Convert to square meters
            total_area_sqm = total_area / 1000000
            
            # Get count of small cuts for this apartment
            small_cuts_count = len(small_tiles_df[small_tiles_df['apartment_name'] == apt_name]) if not small_tiles_df.empty else 0
            
            area_by_apt.append({
                'APPATMENT NO': apt_name,
                'TYPE': 'TL-1',
                'AREA (sqm)': f"{total_area_sqm:.3f}",
                'STATUS': f"Includes {small_cuts_count} small cuts" if small_cuts_count > 0 else "No small cuts"
            })
        
        # Create area dataframe
        area_export = pd.DataFrame(area_by_apt)
        
        # WASTAGE ANALYSIS
        print("\n🔄 Calculating tile wastage percentages...")
        
        # Get the first tile to calculate standard tile dimensions
        sample_tiles = [tile for _, tile in tiles_df.iterrows() 
                        if 'actual_width' in tile and tile['actual_width'] > 0 
                        and 'actual_height' in tile and tile['actual_height'] > 0]
        
        if sample_tiles:
            # Get dimensions from the first valid tile
            standard_tile = sample_tiles[0]
            tile_width = standard_tile.get('actual_width', 0)
            tile_height = standard_tile.get('actual_height', 0)
            standard_tile_area = (tile_width * tile_height) / 1000000  # in square meters
            
            print(f"✅ Using standard tile size: {tile_width}mm x {tile_height}mm = {standard_tile_area:.4f} sqm")
            print(f"✅ Half-tile threshold: {min(tile_width, tile_height)/2}mm")
            
            # Calculate wastage by apartment
            wastage_analysis = []
            
            for apt_name in sorted(all_tiles_for_area['apartment_name'].unique()):
                # Get apartment tiles
                apt_df = all_tiles_for_area[all_tiles_for_area['apartment_name'] == apt_name]
                
                # Calculate actual apartment area
                actual_apt_area = sum(
                    tile['polygon'].area 
                    for _, tile in apt_df.iterrows() 
                    if isinstance(tile['polygon'], (Polygon, MultiPolygon))
                ) / 1000000  # in square meters
                
                # Count tiles by classification
                full_count = len(apt_df[apt_df['classification'] == 'full'])
                irregular_count = len(apt_df[apt_df['classification'] == 'irregular'])
                
                # Handle cut tiles based on size
                cut_tiles = []
                if has_pattern:
                    cut_x = apt_df[apt_df['classification'] == 'cut_x']
                    cut_y = apt_df[apt_df['classification'] == 'cut_y']
                    cut_tiles = pd.concat([cut_x, cut_y])
                else:
                    cut_tiles = apt_df[apt_df['classification'] == 'all_cut']
                
                # For each cut tile, determine if it's more than half or less than half
                half_tile_count = 0
                full_equivalent_count = 0
                
                for _, tile in cut_tiles.iterrows():
                    # Get the cut side dimension
                    if 'cut_side' in tile and tile['cut_side'] is not None:
                        cut_dimension = tile['cut_side']
                        
                        # Compare with standard tile dimensions
                        min_standard_dimension = min(tile_width, tile_height)
                        
                        if cut_dimension > (min_standard_dimension / 2):
                            # Cut side is more than half the standard dimension
                            full_equivalent_count += 1
                        else:
                            # Cut side is less than half the standard dimension
                            half_tile_count += 1
                    elif isinstance(tile['polygon'], (Polygon, MultiPolygon)):
                        # Fallback to area calculation if cut_side is not available
                        minx, miny, maxx, maxy = tile['polygon'].bounds
                        measured_width = maxx - minx
                        measured_height = maxy - miny
                        
                        # Use the smaller dimension for comparison
                        min_dimension = min(measured_width, measured_height)
                        min_standard_dimension = min(tile_width, tile_height)
                        
                        if min_dimension > (min_standard_dimension / 2):
                            full_equivalent_count += 1
                        else:
                            half_tile_count += 1
                
                # Calculate total full-tile equivalents
                total_full_equivalents = (
                    full_count +           # Full tiles
                    irregular_count +      # Irregular tiles count as full
                    full_equivalent_count + # Cut tiles > 50%
                    (half_tile_count / 2)   # Cut tiles < 50% (2 make 1)
                )
                
                # Theoretical area based on perfect coverage (actual apartment area)
                theoretical_area = actual_apt_area
                
                # Area of tiles actually required (in sqm)
                actual_tiles_area = total_full_equivalents * standard_tile_area
                
                # Calculate wastage
                wastage_area = actual_tiles_area - theoretical_area
                wastage_percentage = (wastage_area / theoretical_area) * 100 if theoretical_area > 0 else 0
                
                wastage_analysis.append({
                    'APPATMENT NO': apt_name,
                    'ACTUAL AREA (sqm)': f"{actual_apt_area:.3f}",
                    'FULL TILES': full_count,
                    'IRREGULAR TILES': irregular_count,
                    'CUT TILES >50% DIM': full_equivalent_count,
                    'CUT TILES <50% DIM': half_tile_count,
                    'TOTAL FULL EQUIVALENTS': f"{total_full_equivalents:.1f}",
                    'AREA NEEDED (sqm)': f"{theoretical_area:.3f}",
                    'AREA OF TILES REQUIRED (sqm)': f"{actual_tiles_area:.3f}",
                    'WASTAGE (sqm)': f"{wastage_area:.3f}",
                    'WASTAGE (%)': f"{wastage_percentage:.2f}%"
                })
            
            # Create wastage dataframe
            wastage_export = pd.DataFrame(wastage_analysis)
        else:
            print("⚠️ Could not find standard tile dimensions for wastage calculation")
            # Create empty wastage dataframe with columns
            wastage_export = pd.DataFrame(columns=[
                'APPATMENT NO', 'ACTUAL AREA (sqm)', 'FULL TILES', 'IRREGULAR TILES',
                'CUT TILES >50% DIM', 'CUT TILES <50% DIM', 'TOTAL FULL EQUIVALENTS',
                'AREA NEEDED (sqm)', 'AREA OF TILES REQUIRED (sqm)',
                'WASTAGE (sqm)', 'WASTAGE (%)'
            ])
        
        # Save files
        print("\n💾 Creating single comprehensive final report...")

        # Define single Excel file path
        final_report_path = os.path.join(export_path, f'{output_prefix}_FINAL_REPORT.xlsx')
        
        # Save all data to a single Excel file with multiple sheets
        print(f"Creating comprehensive final report: {final_report_path}")
        with pd.ExcelWriter(final_report_path) as writer:
            # Start with summary and wastage analysis
            wastage_export.to_excel(writer, sheet_name='1. Wastage Analysis', index=False)
            area_export.to_excel(writer, sheet_name='2. Area Summary', index=False)
            
            # Add tiles data
            full_tiles_export.to_excel(writer, sheet_name='3. Full Tiles', index=False)
            
            if has_pattern:
                # Add detailed cut_x and cut_y tiles (with tile sizes)
                cut_x_export.to_excel(writer, sheet_name='4A. Raw Cut X Tiles', index=False)
                cut_y_export.to_excel(writer, sheet_name='5A. Raw Cut Y Tiles', index=False)
                
                # Add simplified cut_x and cut_y tiles (without tile sizes)
                cut_x_simple.to_excel(writer, sheet_name='4B. Cut X Simplified', index=False)
                cut_y_simple.to_excel(writer, sheet_name='5B. Cut Y Simplified', index=False)
            else:
                # Add detailed and simplified all_cut tiles
                all_cut_export.to_excel(writer, sheet_name='4A. Raw All Cut Tiles', index=False)
                all_cut_simple.to_excel(writer, sheet_name='4B. All Cut Simplified', index=False)
            
            # Add small cuts and statistics at the end
            small_cuts_export.to_excel(writer, sheet_name='6. Small Cuts Summary', index=False)
            
            # Add statistics
            stats_data = {
                'Statistic': [
                    'Total Tiles Remaining',
                    'Full Tiles',
                    'Irregular Tiles',
                    'Small Cuts Removed',
                    'Size Threshold'
                ],
                'Value': [
                    len(tiles_df),
                    len(full_df),
                    len(irregular_df),
                    len(small_tiles_df),
                    f"<{size_threshold}mm"
                ]
            }
            
            if has_pattern:
                stats_data['Statistic'].extend(['Cut X Tiles', 'Cut Y Tiles'])
                stats_data['Value'].extend([len(cut_x_df), len(cut_y_df)])
            else:
                stats_data['Statistic'].append('All Cut Tiles')
                stats_data['Value'].append(len(all_cut_df))
            
            pd.DataFrame(stats_data).to_excel(writer, sheet_name='7. Statistics', index=False)

        print("\n📋 Final report saved in the current directory.")

        print("\n✅ Export complete!")
        print(f"   Final report created: {final_report_path}")

        return {
            'full_tiles': full_tiles_export,
            'small_cuts_summary': small_cuts_export,
            'area_summary': area_export,
            'wastage_analysis': wastage_export,
            'export_path': final_report_path,
            **(({'cut_x_tiles': cut_x_export, 'cut_x_simple': cut_x_simple, 
                  'cut_y_tiles': cut_y_export, 'cut_y_simple': cut_y_simple}) 
                if has_pattern else 
                {'all_cut_tiles': all_cut_export, 'all_cut_simple': all_cut_simple})
        }

    def create_apartment_workbook(self, apartment_name, apartment_data, x_inv_data=None, y_inv_data=None, has_pattern=True):
        """Create Excel workbook for a specific apartment"""
        temp_dir = "excel_exports"
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
            
        filename = f"{temp_dir}/{apartment_name}_CutList.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            if has_pattern:
                # Separate X and Y cuts
                x_cuts = apartment_data[apartment_data['Group ID'].str.startswith(('X', 'IX', 'OX'), na=False) | 
                                      (apartment_data['Group ID'] == '') | apartment_data['Group ID'].isna()].copy()
                y_cuts = apartment_data[apartment_data['Group ID'].str.startswith(('Y', 'IY', 'OY'), na=False)].copy()
                
                x_cuts = x_cuts.reset_index(drop=True)
                y_cuts = y_cuts.reset_index(drop=True)
                
                if not x_cuts.empty:
                    x_cuts_for_format = x_cuts.copy()
                    x_export = x_cuts.drop(columns=['Color']) if 'Color' in x_cuts.columns else x_cuts
                    x_export.to_excel(writer, sheet_name='Cut X', index=False)
                    self.apply_excel_formatting(writer.sheets['Cut X'], x_cuts_for_format)
                
                if not y_cuts.empty:
                    y_cuts_for_format = y_cuts.copy()
                    y_export = y_cuts.drop(columns=['Color']) if 'Color' in y_cuts.columns else y_cuts
                    y_export.to_excel(writer, sheet_name='Cut Y', index=False)
                    self.apply_excel_formatting(writer.sheets['Cut Y'], y_cuts_for_format)
            else:
                apartment_data_reset = apartment_data.reset_index(drop=True)
                apartment_data_for_format = apartment_data_reset.copy()
                all_export = apartment_data_reset.drop(columns=['Color']) if 'Color' in apartment_data_reset.columns else apartment_data_reset
                all_export.to_excel(writer, sheet_name='Cut List', index=False)
                self.apply_excel_formatting(writer.sheets['Cut List'], apartment_data_for_format)
            
            # Add inventory sheets
            if x_inv_data is not None and not x_inv_data.empty:
                x_inv_reset = x_inv_data.reset_index(drop=True)
                x_inv_for_format = x_inv_reset.copy()
                x_inv_export = x_inv_reset.drop(columns=['Color']) if 'Color' in x_inv_reset.columns else x_inv_reset
                x_inv_export.to_excel(writer, sheet_name='Inventory X', index=False)
                self.apply_excel_formatting(writer.sheets['Inventory X'], x_inv_for_format)
            
            if y_inv_data is not None and not y_inv_data.empty:
                y_inv_reset = y_inv_data.reset_index(drop=True)
                y_inv_for_format = y_inv_reset.copy()
                y_inv_export = y_inv_reset.drop(columns=['Color']) if 'Color' in y_inv_reset.columns else y_inv_reset
                y_inv_export.to_excel(writer, sheet_name='Inventory Y', index=False)
                self.apply_excel_formatting(writer.sheets['Inventory Y'], y_inv_for_format)
        
        return filename

    def apply_excel_formatting(self, worksheet, df):
        """Apply background colors to Excel cells based on the Color column"""
        try:
            from openpyxl.styles import PatternFill
            
            color_col_idx = None
            group_id_col_idx = None
            cut_size_col_idx = None
            remaining_size_col_idx = None
            
            for idx, col in enumerate(df.columns):
                if col == 'Color':
                    color_col_idx = idx
                elif col == 'Group ID':
                    group_id_col_idx = idx
                elif col == 'Cut Size':
                    cut_size_col_idx = idx
                elif col == 'Remaining Size':
                    remaining_size_col_idx = idx
            
            for row_idx, row_data in df.iterrows():
                excel_row = row_idx + 2  # +2 for Excel 1-indexing and header
                
                color_value = row_data.get('Color', '')
                if color_value and color_value != '':
                    hex_color = color_value.replace('#', '') if color_value.startswith('#') else color_value
                    
                    if len(hex_color) == 3:
                        hex_color = ''.join([c*2 for c in hex_color])
                    
                    if len(hex_color) == 6:
                        try:
                            int(hex_color, 16)  # Validate hex
                            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                            
                            # Always color Group ID column
                            if group_id_col_idx is not None:
                                worksheet.cell(row=excel_row, column=group_id_col_idx + 1).fill = fill
                            
                            # Color Cut Size or Remaining Size based on conditions
                            piece_type = row_data.get('Piece Type', '')
                            match_type = row_data.get('Match Type', '')
                            status = row_data.get('Status', '')
                            group_id = str(row_data.get('Group ID', ''))
                            
                            is_matched = group_id and group_id != '' and not group_id.startswith('O')
                            
                            if match_type == 'Apartment' and is_matched:
                                if piece_type == 'More than Half' and cut_size_col_idx is not None:
                                    worksheet.cell(row=excel_row, column=cut_size_col_idx + 1).fill = fill
                                elif piece_type == 'Less than Half' and remaining_size_col_idx is not None:
                                    worksheet.cell(row=excel_row, column=remaining_size_col_idx + 1).fill = fill
                            elif status == 'Matched' and cut_size_col_idx is not None:
                                worksheet.cell(row=excel_row, column=cut_size_col_idx + 1).fill = fill
                        
                        except (ValueError, Exception):
                            continue  # Skip invalid colors
        
        except ImportError:
            print("   ⚠️ openpyxl not available - Excel files will be created without colors")
        except Exception as e:
            print(f"   ⚠️ Formatting error: {e}")

    def create_cut_pieces_summary(self, tiles_df, has_pattern, tile_width, tile_height):
        """Create cut pieces summary for matching process"""
        print("\n📊 Creating cut pieces summary...")
        
        # Calculate half-tile threshold
        half_threshold = min(tile_width, tile_height) / 2
        
        if has_pattern:
            # Process X-direction cuts
            cut_x_tiles = tiles_df[tiles_df['classification'] == 'cut_x'].copy()
            x_less_than_half = []
            x_more_than_half = []
            
            for _, tile in cut_x_tiles.iterrows():
                cut_size = tile['cut_side']
                remaining_size = tile_width - cut_size
                
                summary = {
                    'Apartment': tile['apartment_name'],
                    'Location': tile['room_name'],
                    'Cut Size (mm)': round(cut_size, 1),
                    'Remaining Size (mm)': round(remaining_size, 1),
                    'Count': 1
                }
                
                if cut_size < half_threshold:
                    x_less_than_half.append(summary)
                else:
                    x_more_than_half.append(summary)
            
            # Process Y-direction cuts
            cut_y_tiles = tiles_df[tiles_df['classification'] == 'cut_y'].copy()
            y_less_than_half = []
            y_more_than_half = []
            
            for _, tile in cut_y_tiles.iterrows():
                cut_size = tile['cut_side']
                remaining_size = tile_height - cut_size
                
                summary = {
                    'Apartment': tile['apartment_name'],
                    'Location': tile['room_name'],
                    'Cut Size (mm)': round(cut_size, 1),
                    'Remaining Size (mm)': round(remaining_size, 1),
                    'Count': 1
                }
                
                if cut_size < half_threshold:
                    y_less_than_half.append(summary)
                else:
                    y_more_than_half.append(summary)
            
            # Convert to DataFrames and aggregate
            def aggregate_summary(summary_list):
                if not summary_list:
                    return pd.DataFrame()
                df = pd.DataFrame(summary_list)
                return df.groupby(['Apartment', 'Location', 'Cut Size (mm)', 'Remaining Size (mm)']).sum().reset_index()
            
            x_less_than_half_df = aggregate_summary(x_less_than_half)
            x_more_than_half_df = aggregate_summary(x_more_than_half)
            y_less_than_half_df = aggregate_summary(y_less_than_half)
            y_more_than_half_df = aggregate_summary(y_more_than_half)
            
            return {
                'x_less_than_half': x_less_than_half_df,
                'x_more_than_half': x_more_than_half_df,
                'y_less_than_half': y_less_than_half_df,
                'y_more_than_half': y_more_than_half_df,
                'has_pattern': True,
                'tile_width': tile_width,
                'tile_height': tile_height,
                'half_threshold': half_threshold
            }
            
        else:
            # Process all cuts (no pattern)
            all_cut_tiles = tiles_df[tiles_df['classification'] == 'all_cut'].copy()
            all_less_than_half = []
            all_more_than_half = []
            
            for _, tile in all_cut_tiles.iterrows():
                cut_size = tile['cut_side']
                # For no pattern, use minimum dimension for remaining size calculation
                min_dimension = min(tile_width, tile_height)
                remaining_size = min_dimension - cut_size
                
                summary = {
                    'Apartment': tile['apartment_name'],
                    'Location': tile['room_name'],
                    'Cut Size (mm)': round(cut_size, 1),
                    'Remaining Size (mm)': round(remaining_size, 1),
                    'Count': 1
                }
                
                if cut_size < half_threshold:
                    all_less_than_half.append(summary)
                else:
                    all_more_than_half.append(summary)
            
            # Convert to DataFrames and aggregate
            def aggregate_summary(summary_list):
                if not summary_list:
                    return pd.DataFrame()
                df = pd.DataFrame(summary_list)
                return df.groupby(['Apartment', 'Location', 'Cut Size (mm)', 'Remaining Size (mm)']).sum().reset_index()
            
            all_less_than_half_df = aggregate_summary(all_less_than_half)
            all_more_than_half_df = aggregate_summary(all_more_than_half)
            
            return {
                'all_less_than_half': all_less_than_half_df,
                'all_more_than_half': all_more_than_half_df,
                'has_pattern': False,
                'tile_width': tile_width,
                'tile_height': tile_height,
                'half_threshold': half_threshold
            }

    def create_inventory_from_more_than_half(self, cut_pieces_summary):
        """Create inventory from more-than-half pieces"""
        print("\n📦 Creating inventory from more-than-half pieces...")
        
        has_pattern = cut_pieces_summary['has_pattern']
        
        if has_pattern:
            # Process X and Y separately
            x_more = cut_pieces_summary.get('x_more_than_half', pd.DataFrame())
            y_more = cut_pieces_summary.get('y_more_than_half', pd.DataFrame())
            
            # Create inventory summaries
            x_inv_less = pd.DataFrame()
            x_inv_more = pd.DataFrame()
            y_inv_less = pd.DataFrame()
            y_inv_more = pd.DataFrame()
            
            if not x_more.empty:
                x_inv_summary = []
                for _, row in x_more.iterrows():
                    x_inv_summary.append({
                        'Location': f"{row['Apartment']}-{row['Location']}",
                        'Cut Size (mm)': row['Cut Size (mm)'],
                        'Remaining Size (mm)': row['Remaining Size (mm)'],
                        'Count': row['Count']
                    })
                x_inv_df = pd.DataFrame(x_inv_summary)
                
                # Split into less/more based on remaining size
                half_threshold = cut_pieces_summary['half_threshold']
                x_inv_less = x_inv_df[x_inv_df['Remaining Size (mm)'] < half_threshold].copy()
                x_inv_more = x_inv_df[x_inv_df['Remaining Size (mm)'] >= half_threshold].copy()
            
            if not y_more.empty:
                y_inv_summary = []
                for _, row in y_more.iterrows():
                    y_inv_summary.append({
                        'Location': f"{row['Apartment']}-{row['Location']}",
                        'Cut Size (mm)': row['Cut Size (mm)'],
                        'Remaining Size (mm)': row['Remaining Size (mm)'],
                        'Count': row['Count']
                    })
                y_inv_df = pd.DataFrame(y_inv_summary)
                
                # Split into less/more based on remaining size
                half_threshold = cut_pieces_summary['half_threshold']
                y_inv_less = y_inv_df[y_inv_df['Remaining Size (mm)'] < half_threshold].copy()
                y_inv_more = y_inv_df[y_inv_df['Remaining Size (mm)'] >= half_threshold].copy()
            
            # Update the summary
            cut_pieces_summary['x_inv_less_than_half'] = x_inv_less
            cut_pieces_summary['x_inv_more_than_half'] = x_inv_more
            cut_pieces_summary['y_inv_less_than_half'] = y_inv_less
            cut_pieces_summary['y_inv_more_than_half'] = y_inv_more
            
        else:
            # Process all cuts together
            all_more = cut_pieces_summary.get('all_more_than_half', pd.DataFrame())
            
            all_inv_less = pd.DataFrame()
            all_inv_more = pd.DataFrame()
            
            if not all_more.empty:
                all_inv_summary = []
                for _, row in all_more.iterrows():
                    all_inv_summary.append({
                        'Location': f"{row['Apartment']}-{row['Location']}",
                        'Cut Size (mm)': row['Cut Size (mm)'],
                        'Remaining Size (mm)': row['Remaining Size (mm)'],
                        'Count': row['Count']
                    })
                all_inv_df = pd.DataFrame(all_inv_summary)
                
                # Split into less/more based on remaining size
                half_threshold = cut_pieces_summary['half_threshold']
                all_inv_less = all_inv_df[all_inv_df['Remaining Size (mm)'] < half_threshold].copy()
                all_inv_more = all_inv_df[all_inv_df['Remaining Size (mm)'] >= half_threshold].copy()
            
            # Update the summary
            cut_pieces_summary['all_inv_less_than_half'] = all_inv_less
            cut_pieces_summary['all_inv_more_than_half'] = all_inv_more
        
        return cut_pieces_summary


    def create_summary_excel_export(self, tiles_df, small_tiles_df, final_room_df, 
                               size_threshold=10, output_prefix="final_tiles_export", 
                               selected_matching=None):
        """Create a comprehensive summary Excel export with correct cut piece equivalency"""
        
        print("\n📊 Creating Enhanced Summary Excel Export...")
        
        export_path = os.getcwd()
        summary_file_path = os.path.join(export_path, f'{output_prefix}_SUMMARY_REPORT.xlsx')
        
        # Get tile dimensions from sample tile
        sample_tiles = [tile for _, tile in tiles_df.iterrows() 
                    if 'actual_width' in tile and tile['actual_width'] > 0 
                    and 'actual_height' in tile and tile['actual_height'] > 0]
        
        if not sample_tiles:
            print("⚠️ Could not find valid tile dimensions")
            return None
        
        # Standard tile properties
        standard_tile = sample_tiles[0]
        tile_length = standard_tile.get('actual_width', 600)  # mm
        tile_width = standard_tile.get('actual_height', 600)  # mm
        tile_area_mm2 = tile_length * tile_width
        tile_area_m2 = tile_area_mm2 / 1000000  # Convert to square meters
        half_threshold = min(tile_length, tile_width) / 2
        
        print(f"✅ Using tile size: {tile_length}mm × {tile_width}mm = {tile_area_m2:.6f} m²")
        print(f"✅ Half threshold: {half_threshold}mm")
        
        # Get detailed matching data if available
        matching_data = {}
        if selected_matching:
            apartment_summaries = selected_matching.get('apartment_summaries', {})
            clean_tables = selected_matching.get('clean_tables', {})
            
            # Process clean tables to get detailed breakdown
            for apt_name, summary in apartment_summaries.items():
                # Initialize counts
                apt_data = {
                    'total_matched': summary.get('matched_pieces', 0),
                    'total_unmatched': summary.get('unmatched_pieces', 0),
                    'matched_less_than_half': 0,
                    'matched_more_than_half': 0,
                    'unmatched_less_than_half': 0,
                    'unmatched_more_than_half': 0
                }
                
                # Analyze clean tables to get size breakdowns
                for table_name, table_data in clean_tables.items():
                    if isinstance(table_data, list) and len(table_data) > 0:
                        df = pd.DataFrame(table_data)
                        if 'Apartment' in df.columns and 'Cut Size' in df.columns:
                            apt_df = df[df['Apartment'] == apt_name]
                            
                            for _, row in apt_df.iterrows():
                                cut_size = row.get('Cut Size', 0)
                                count = row.get('Count', 0)
                                group_id = row.get('Group ID', '')
                                
                                # Determine if matched or unmatched
                                is_matched = bool(group_id and group_id.strip())
                                
                                # Determine if less than or more than half
                                is_less_than_half = cut_size < half_threshold
                                
                                if is_matched:
                                    if is_less_than_half:
                                        apt_data['matched_less_than_half'] += count
                                    else:
                                        apt_data['matched_more_than_half'] += count
                                else:
                                    if is_less_than_half:
                                        apt_data['unmatched_less_than_half'] += count
                                    else:
                                        apt_data['unmatched_more_than_half'] += count
                
                matching_data[apt_name] = apt_data
        
        # Create the main summary table
        summary_data = []
        
        # Get all apartment data
        all_tiles_for_area = pd.concat([tiles_df, small_tiles_df], ignore_index=True) if not small_tiles_df.empty else tiles_df
        
        for apt_name in sorted(all_tiles_for_area['apartment_name'].unique()):
            apt_tiles = all_tiles_for_area[all_tiles_for_area['apartment_name'] == apt_name]
            
            # Calculate apartment area (from room polygons)
            apartment_area_m2 = 0
            apt_rooms = final_room_df[final_room_df['apartment_name'] == apt_name] if not final_room_df.empty else pd.DataFrame()
            
            if not apt_rooms.empty:
                for _, room in apt_rooms.iterrows():
                    if hasattr(room, 'polygon') and room['polygon'] is not None:
                        apartment_area_m2 += room['polygon'].area / 1000000  # Convert mm² to m²
            else:
                # Fallback: calculate from tile polygons
                apartment_area_m2 = sum(
                    tile['polygon'].area for _, tile in apt_tiles.iterrows() 
                    if hasattr(tile, 'polygon') and tile['polygon'] is not None
                ) / 1000000
            
            # Count different tile types
            full_tiles = len(apt_tiles[apt_tiles['classification'] == 'full'])
            irregular_tiles = len(apt_tiles[apt_tiles['classification'] == 'irregular'])
            
            # Get detailed matching data
            if apt_name in matching_data:
                apt_match_data = matching_data[apt_name]
                matched_less_half = apt_match_data['matched_less_than_half']
                matched_more_half = apt_match_data['matched_more_than_half']
                unmatched_less_half = apt_match_data['unmatched_less_than_half']
                unmatched_more_half = apt_match_data['unmatched_more_than_half']
            else:
                # If no matching data available, assume all are unmatched and split evenly
                cut_tiles = apt_tiles[apt_tiles['classification'].isin(['cut_x', 'cut_y', 'all_cut'])]
                total_cut_tiles = len(cut_tiles)
                matched_less_half = 0
                matched_more_half = 0
                unmatched_less_half = total_cut_tiles // 2  # Rough estimate
                unmatched_more_half = total_cut_tiles - unmatched_less_half
            
            # Calculate full tile equivalents using the CORRECT formula:
            # FULL + IRREGULAR + (MATCHED)/2 + (UNMATCHED_LESS_THAN_HALF)/2 + (UNMATCHED_MORE_THAN_HALF)
            matched_equivalent = (matched_less_half + matched_more_half) / 2
            unmatched_less_equivalent = unmatched_less_half / 2
            unmatched_more_equivalent = unmatched_more_half  # Full equivalent
            
            total_full_equivalents = (full_tiles + irregular_tiles + 
                                    matched_equivalent + 
                                    unmatched_less_equivalent + 
                                    unmatched_more_equivalent)
            
            # Calculate tiling area (total tiles × tile size)
            tiling_area_m2 = total_full_equivalents * tile_area_m2
            
            # Calculate wastage percentage
            wastage_percentage = ((tiling_area_m2 - apartment_area_m2) / apartment_area_m2) * 100 if apartment_area_m2 > 0 else 0
            
            # Create the summary row showing the breakdown
            total_matched = matched_less_half + matched_more_half
            total_unmatched = unmatched_less_half + unmatched_more_half
            
            summary_data.append({
                'APARTMENT NO.': apt_name,
                'APARTMENT AREA': f"{apartment_area_m2:.1f}",
                'FULL TILES': f"({full_tiles} + {irregular_tiles} + ({total_matched})/2 + ({unmatched_less_half})/2 + {unmatched_more_half}) = {total_full_equivalents:.1f}",
                'TILING AREA': f"{total_full_equivalents:.1f} * {tile_area_m2:.6f} = {tiling_area_m2:.3f}",
                'WASTAGE %': f"({tiling_area_m2:.3f} - {apartment_area_m2:.1f}) / {apartment_area_m2:.1f} = {wastage_percentage:.2f}%"
            })
        
        # Convert to DataFrame
        summary_df = pd.DataFrame(summary_data)
        
        # Create additional detailed breakdown table
        detailed_breakdown = []
        
        for apt_name in sorted(all_tiles_for_area['apartment_name'].unique()):
            apt_tiles = all_tiles_for_area[all_tiles_for_area['apartment_name'] == apt_name]
            
            # Calculate apartment area
            apartment_area_m2 = 0
            apt_rooms = final_room_df[final_room_df['apartment_name'] == apt_name] if not final_room_df.empty else pd.DataFrame()
            
            if not apt_rooms.empty:
                for _, room in apt_rooms.iterrows():
                    if hasattr(room, 'polygon') and room['polygon'] is not None:
                        apartment_area_m2 += room['polygon'].area / 1000000
            else:
                apartment_area_m2 = sum(
                    tile['polygon'].area for _, tile in apt_tiles.iterrows() 
                    if hasattr(tile, 'polygon') and tile['polygon'] is not None
                ) / 1000000
            
            # Detailed counts
            full_count = len(apt_tiles[apt_tiles['classification'] == 'full'])
            irregular_count = len(apt_tiles[apt_tiles['classification'] == 'irregular'])
            cut_x_count = len(apt_tiles[apt_tiles['classification'] == 'cut_x'])
            cut_y_count = len(apt_tiles[apt_tiles['classification'] == 'cut_y'])
            all_cut_count = len(apt_tiles[apt_tiles['classification'] == 'all_cut'])
            small_cuts_count = len(small_tiles_df[small_tiles_df['apartment_name'] == apt_name]) if not small_tiles_df.empty else 0
            
            total_cuts = cut_x_count + cut_y_count + all_cut_count
            total_tiles = full_count + irregular_count + total_cuts
            
            # Get detailed matching data
            if apt_name in matching_data:
                apt_match_data = matching_data[apt_name]
                matched_less_half = apt_match_data['matched_less_than_half']
                matched_more_half = apt_match_data['matched_more_than_half']
                unmatched_less_half = apt_match_data['unmatched_less_than_half']
                unmatched_more_half = apt_match_data['unmatched_more_than_half']
            else:
                matched_less_half = 0
                matched_more_half = 0
                unmatched_less_half = total_cuts // 2
                unmatched_more_half = total_cuts - unmatched_less_half
            
            # Calculate using the correct formula
            matched_equivalent = (matched_less_half + matched_more_half) / 2
            unmatched_less_equivalent = unmatched_less_half / 2
            unmatched_more_equivalent = unmatched_more_half
            
            total_full_equivalents = (full_count + irregular_count + 
                                    matched_equivalent + 
                                    unmatched_less_equivalent + 
                                    unmatched_more_equivalent)
            
            tiling_area_m2 = total_full_equivalents * tile_area_m2
            wastage_area_m2 = tiling_area_m2 - apartment_area_m2
            wastage_percentage = (wastage_area_m2 / apartment_area_m2) * 100 if apartment_area_m2 > 0 else 0
            
            detailed_breakdown.append({
                'APARTMENT NO.': apt_name,
                'APARTMENT AREA (m²)': round(apartment_area_m2, 2),
                'FULL TILES': full_count,
                'IRREGULAR TILES': irregular_count,
                'CUT X TILES': cut_x_count,
                'CUT Y TILES': cut_y_count,
                'ALL CUT TILES': all_cut_count,
                'TOTAL CUT TILES': total_cuts,
                'MATCHED < HALF': matched_less_half,
                'MATCHED > HALF': matched_more_half,
                'UNMATCHED < HALF': unmatched_less_half,
                'UNMATCHED > HALF': unmatched_more_half,
                'MATCHED EQUIVALENT': round(matched_equivalent, 1),
                'UNMATCHED < HALF EQUIVALENT': round(unmatched_less_equivalent, 1),
                'UNMATCHED > HALF EQUIVALENT': round(unmatched_more_equivalent, 1),
                'SMALL CUTS (<{}mm)'.format(size_threshold): small_cuts_count,
                'TOTAL TILES': total_tiles,
                'FULL TILE EQUIVALENTS': round(total_full_equivalents, 1),
                'TILE SIZE (m²)': tile_area_m2,
                'TILING AREA (m²)': round(tiling_area_m2, 3),
                'WASTAGE AREA (m²)': round(wastage_area_m2, 3),
                'WASTAGE %': round(wastage_percentage, 2)
            })
        
        detailed_df = pd.DataFrame(detailed_breakdown)
        
        # Create tile specifications table
        tile_specs_data = [{
            'PARAMETER': 'TILE LENGTH (mm)',
            'VALUE': tile_length
        }, {
            'PARAMETER': 'TILE WIDTH (mm)', 
            'VALUE': tile_width
        }, {
            'PARAMETER': 'TILE AREA (mm²)',
            'VALUE': tile_area_mm2
        }, {
            'PARAMETER': 'TILE AREA (m²)',
            'VALUE': f"{tile_area_m2:.6f}"
        }, {
            'PARAMETER': 'HALF THRESHOLD (mm)',
            'VALUE': f"{half_threshold:.1f}"
        }, {
            'PARAMETER': 'SMALL CUT THRESHOLD (mm)',
            'VALUE': size_threshold
        }]
        
        tile_specs_df = pd.DataFrame(tile_specs_data)
        
        # Create project summary
        total_apartments = len(summary_data)
        total_apartment_area = sum(float(row['APARTMENT AREA']) for row in summary_data)
        total_tiles_all = sum(row['TOTAL TILES'] for row in detailed_breakdown)
        total_full_equiv = sum(row['FULL TILE EQUIVALENTS'] for row in detailed_breakdown)
        total_tiling_area = total_full_equiv * tile_area_m2
        overall_wastage = ((total_tiling_area - total_apartment_area) / total_apartment_area) * 100 if total_apartment_area > 0 else 0
        
        project_summary_data = [{
            'METRIC': 'TOTAL APARTMENTS',
            'VALUE': total_apartments
        }, {
            'METRIC': 'TOTAL APARTMENT AREA (m²)',
            'VALUE': round(total_apartment_area, 2)
        }, {
            'METRIC': 'TOTAL TILES',
            'VALUE': total_tiles_all
        }, {
            'METRIC': 'TOTAL FULL TILE EQUIVALENTS',
            'VALUE': round(total_full_equiv, 1)
        }, {
            'METRIC': 'TOTAL TILING AREA (m²)',
            'VALUE': round(total_tiling_area, 3)
        }, {
            'METRIC': 'OVERALL WASTAGE %',
            'VALUE': f"{round(overall_wastage, 2)}%"
        }]
        
        project_summary_df = pd.DataFrame(project_summary_data)
        
        # Save to Excel with multiple sheets
        print(f"💾 Creating summary report: {summary_file_path}")
        
        with pd.ExcelWriter(summary_file_path, engine='openpyxl') as writer:
            # Sheet 1: Main Summary (with corrected formula breakdown)
            summary_df.to_excel(writer, sheet_name='1. Summary Report', index=False)
            
            # Sheet 2: Detailed Breakdown
            detailed_df.to_excel(writer, sheet_name='2. Detailed Breakdown', index=False)
            
            # Sheet 3: Project Summary
            project_summary_df.to_excel(writer, sheet_name='3. Project Summary', index=False)
            
            # Sheet 4: Tile Specifications
            tile_specs_df.to_excel(writer, sheet_name='4. Tile Specifications', index=False)
            
            # Apply formatting to make it look professional
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 60)  # Increased for longer formulas
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Bold headers
                for cell in worksheet[1]:
                    cell.font = cell.font.copy(bold=True)
        
        print(f"✅ Summary report created: {summary_file_path}")
        print("\n📋 Report includes:")
        print("   • Sheet 1: Summary Report (corrected equivalency formula)")
        print("   • Sheet 2: Detailed Breakdown (with cut piece analysis)")
        print("   • Sheet 3: Project Summary (totals)")
        print("   • Sheet 4: Tile Specifications")
        
        return {
            'summary_file_path': summary_file_path,
            'summary_data': summary_df,
            'detailed_breakdown': detailed_df,
            'project_summary': project_summary_df,
            'tile_specs': tile_specs_df
        }

    # Also add this method to integrate with your existing export function
    def export_remaining_tiles_with_enhanced_summary(self, tiles_df, small_tiles_df, final_room_df, 
                                                size_threshold=10, output_prefix="final_tiles_export"):
        """Enhanced export that includes both the original report and the new summary"""
        
        # First create the original comprehensive report
        original_result = self.export_remaining_tiles_with_wastage_analysis(
            tiles_df, small_tiles_df, final_room_df, size_threshold, output_prefix
        )
        
        # Then create the enhanced summary report
        summary_result = self.create_summary_excel_export(
            tiles_df, small_tiles_df, final_room_df, size_threshold, output_prefix
        )
        
        print("\n✅ Export complete with enhanced summary!")
        print(f"   📊 Original comprehensive report: {original_result.get('export_path', 'Not created')}")
        print(f"   📋 Enhanced summary report: {summary_result.get('summary_file_path', 'Not created')}")
        
        return {
            'original_report': original_result,
            'summary_report': summary_result,
            'export_complete': True
        }
    
    def create_enhanced_summary_data_for_master(self, tiles_df, small_tiles_df, final_room_df, selected_matching):
        """Create enhanced summary data specifically for master workbook integration"""
        try:
            # Get tile dimensions from sample tile
            sample_tiles = [tile for _, tile in tiles_df.iterrows() 
                           if 'actual_width' in tile and tile['actual_width'] > 0 
                           and 'actual_height' in tile and tile['actual_height'] > 0]
            
            if not sample_tiles:
                return None
            
            # Standard tile properties
            standard_tile = sample_tiles[0]
            tile_length = standard_tile.get('actual_width', 600)  # mm
            tile_width = standard_tile.get('actual_height', 600)  # mm
            tile_area_mm2 = tile_length * tile_width
            tile_area_m2 = tile_area_mm2 / 1000000  # Convert to square meters
            half_threshold = min(tile_length, tile_width) / 2
            
            # Get detailed matching data if available
            matching_data = {}
            if selected_matching:
                apartment_summaries = selected_matching.get('apartment_summaries', {})
                clean_tables = selected_matching.get('clean_tables', {})
                
                # Process clean tables to get detailed breakdown
                for apt_name, summary in apartment_summaries.items():
                    # Initialize counts
                    apt_data = {
                        'total_matched': summary.get('matched_pieces', 0),
                        'total_unmatched': summary.get('unmatched_pieces', 0),
                        'matched_less_than_half': 0,
                        'matched_more_than_half': 0,
                        'unmatched_less_than_half': 0,
                        'unmatched_more_than_half': 0
                    }
                    
                    # Analyze clean tables to get size breakdowns
                    for table_name, table_data in clean_tables.items():
                        if isinstance(table_data, list) and len(table_data) > 0:
                            df = pd.DataFrame(table_data)
                            if 'Apartment' in df.columns and 'Cut Size' in df.columns:
                                apt_df = df[df['Apartment'] == apt_name]
                                
                                for _, row in apt_df.iterrows():
                                    cut_size = row.get('Cut Size', 0)
                                    count = row.get('Count', 0)
                                    group_id = row.get('Group ID', '')
                                    
                                    # Determine if matched or unmatched
                                    is_matched = bool(group_id and group_id.strip())
                                    
                                    # Determine if less than or more than half
                                    is_less_than_half = cut_size < half_threshold
                                    
                                    if is_matched:
                                        if is_less_than_half:
                                            apt_data['matched_less_than_half'] += count
                                        else:
                                            apt_data['matched_more_than_half'] += count
                                    else:
                                        if is_less_than_half:
                                            apt_data['unmatched_less_than_half'] += count
                                        else:
                                            apt_data['unmatched_more_than_half'] += count
                    
                    matching_data[apt_name] = apt_data
            
            # Create the main summary table
            summary_data = []
            all_tiles_for_area = pd.concat([tiles_df, small_tiles_df], ignore_index=True) if not small_tiles_df.empty else tiles_df
            
            for apt_name in sorted(all_tiles_for_area['apartment_name'].unique()):
                apt_tiles = all_tiles_for_area[all_tiles_for_area['apartment_name'] == apt_name]
                
                # Calculate apartment area (from room polygons)
                apartment_area_m2 = 0
                apt_rooms = final_room_df[final_room_df['apartment_name'] == apt_name] if not final_room_df.empty else pd.DataFrame()
                
                if not apt_rooms.empty:
                    for _, room in apt_rooms.iterrows():
                        if hasattr(room, 'polygon') and room['polygon'] is not None:
                            apartment_area_m2 += room['polygon'].area / 1000000  # Convert mm² to m²
                else:
                    # Fallback: calculate from tile polygons
                    apartment_area_m2 = sum(
                        tile['polygon'].area for _, tile in apt_tiles.iterrows() 
                        if hasattr(tile, 'polygon') and tile['polygon'] is not None
                    ) / 1000000
                
                # Count different tile types
                full_tiles = len(apt_tiles[apt_tiles['classification'] == 'full'])
                irregular_tiles = len(apt_tiles[apt_tiles['classification'] == 'irregular'])
                
                # Get detailed matching data
                if apt_name in matching_data:
                    apt_match_data = matching_data[apt_name]
                    matched_less_half = apt_match_data['matched_less_than_half']
                    matched_more_half = apt_match_data['matched_more_than_half']
                    unmatched_less_half = apt_match_data['unmatched_less_than_half']
                    unmatched_more_half = apt_match_data['unmatched_more_than_half']
                else:
                    # If no matching data available, assume all are unmatched and split evenly
                    cut_tiles = apt_tiles[apt_tiles['classification'].isin(['cut_x', 'cut_y', 'all_cut'])]
                    total_cut_tiles = len(cut_tiles)
                    matched_less_half = 0
                    matched_more_half = 0
                    unmatched_less_half = total_cut_tiles // 2  # Rough estimate
                    unmatched_more_half = total_cut_tiles - unmatched_less_half
                
                # *** DESIGN WASTAGE CALCULATION (Before Optimization) ***
                # Design treats ALL cuts as unmatched (no optimization)
                all_less_half = matched_less_half + unmatched_less_half
                all_more_half = matched_more_half + unmatched_more_half
                
                design_total_equivalents = (full_tiles + irregular_tiles + 
                                          (all_less_half / 2) + all_more_half)
                design_tiling_area_m2 = design_total_equivalents * tile_area_m2
                design_wastage_percentage = ((design_tiling_area_m2 - apartment_area_m2) / apartment_area_m2) * 100 if apartment_area_m2 > 0 else 0
                
                # *** OPTIMISED WASTAGE CALCULATION (After Optimization) ***
                # Calculate full tile equivalents using the CORRECT formula:
                # FULL + IRREGULAR + (MATCHED)/2 + (UNMATCHED_LESS_THAN_HALF)/2 + (UNMATCHED_MORE_THAN_HALF)
                matched_equivalent = (matched_less_half + matched_more_half) / 2
                unmatched_less_equivalent = unmatched_less_half / 2
                unmatched_more_equivalent = unmatched_more_half  # Full equivalent
                
                optimised_total_equivalents = (full_tiles + irregular_tiles + 
                                             matched_equivalent + 
                                             unmatched_less_equivalent + 
                                             unmatched_more_equivalent)
                
                # Calculate tiling area (total tiles × tile size)
                optimised_tiling_area_m2 = optimised_total_equivalents * tile_area_m2
                
                # Calculate wastage percentage
                optimised_wastage_percentage = ((optimised_tiling_area_m2 - apartment_area_m2) / apartment_area_m2) * 100 if apartment_area_m2 > 0 else 0
                
                # Calculate savings
                wastage_savings = design_wastage_percentage - optimised_wastage_percentage
                
                # Create the summary row showing the breakdown
                total_matched = matched_less_half + matched_more_half
                total_unmatched = unmatched_less_half + unmatched_more_half
                
                summary_data.append({
                    'APARTMENT NO.': apt_name,
                    'APARTMENT AREA (m²)': round(apartment_area_m2, 2),
                    'FULL TILES': full_tiles,
                    'IRREGULAR TILES': irregular_tiles,
                    'MATCHED CUTS': total_matched,
                    'UNMATCHED CUTS': total_unmatched,
                    'DESIGN EQUIVALENTS': round(design_total_equivalents, 1),
                    'OPTIMISED EQUIVALENTS': round(optimised_total_equivalents, 1),
                    'DESIGN TILING AREA (m²)': round(design_tiling_area_m2, 3),
                    'OPTIMISED TILING AREA (m²)': round(optimised_tiling_area_m2, 3),
                    'DESIGN WASTAGE %': round(design_wastage_percentage, 2),
                    'OPTIMISED WASTAGE %': round(optimised_wastage_percentage, 2),
                    'SAVINGS %': round(wastage_savings, 2)
                })
            
            # Convert to DataFrame
            summary_df = pd.DataFrame(summary_data)
            
            # Create additional detailed breakdown table
            detailed_breakdown = []
            for apt_name in sorted(all_tiles_for_area['apartment_name'].unique()):
                apt_tiles = all_tiles_for_area[all_tiles_for_area['apartment_name'] == apt_name]
                
                # Calculate apartment area
                apartment_area_m2 = 0
                apt_rooms = final_room_df[final_room_df['apartment_name'] == apt_name] if not final_room_df.empty else pd.DataFrame()
                
                if not apt_rooms.empty:
                    for _, room in apt_rooms.iterrows():
                        if hasattr(room, 'polygon') and room['polygon'] is not None:
                            apartment_area_m2 += room['polygon'].area / 1000000
                else:
                    apartment_area_m2 = sum(
                        tile['polygon'].area for _, tile in apt_tiles.iterrows() 
                        if hasattr(tile, 'polygon') and tile['polygon'] is not None
                    ) / 1000000
                
                # Detailed counts
                full_count = len(apt_tiles[apt_tiles['classification'] == 'full'])
                irregular_count = len(apt_tiles[apt_tiles['classification'] == 'irregular'])
                cut_x_count = len(apt_tiles[apt_tiles['classification'] == 'cut_x'])
                cut_y_count = len(apt_tiles[apt_tiles['classification'] == 'cut_y'])
                all_cut_count = len(apt_tiles[apt_tiles['classification'] == 'all_cut'])
                small_cuts_count = len(small_tiles_df[small_tiles_df['apartment_name'] == apt_name]) if not small_tiles_df.empty else 0
                
                total_cuts = cut_x_count + cut_y_count + all_cut_count
                total_tiles = full_count + irregular_count + total_cuts
                
                # Get detailed matching data
                if apt_name in matching_data:
                    apt_match_data = matching_data[apt_name]
                    matched_less_half = apt_match_data['matched_less_than_half']
                    matched_more_half = apt_match_data['matched_more_than_half']
                    unmatched_less_half = apt_match_data['unmatched_less_than_half']
                    unmatched_more_half = apt_match_data['unmatched_more_than_half']
                else:
                    matched_less_half = 0
                    matched_more_half = 0
                    unmatched_less_half = total_cuts // 2
                    unmatched_more_half = total_cuts - unmatched_less_half
                
                # Calculate using both formulas
                # Design (before optimization)
                all_less_half = matched_less_half + unmatched_less_half
                all_more_half = matched_more_half + unmatched_more_half
                design_total_equivalents = (full_count + irregular_count + 
                                          (all_less_half / 2) + all_more_half)
                design_tiling_area_m2 = design_total_equivalents * tile_area_m2
                design_wastage_percentage = ((design_tiling_area_m2 - apartment_area_m2) / apartment_area_m2) * 100 if apartment_area_m2 > 0 else 0
                
                # Optimised (after optimization)
                matched_equivalent = (matched_less_half + matched_more_half) / 2
                unmatched_less_equivalent = unmatched_less_half / 2
                unmatched_more_equivalent = unmatched_more_half
                
                optimised_total_equivalents = (full_count + irregular_count + 
                                             matched_equivalent + 
                                             unmatched_less_equivalent + 
                                             unmatched_more_equivalent)
                
                optimised_tiling_area_m2 = optimised_total_equivalents * tile_area_m2
                optimised_wastage_percentage = ((optimised_tiling_area_m2 - apartment_area_m2) / apartment_area_m2) * 100 if apartment_area_m2 > 0 else 0
                
                detailed_breakdown.append({
                    'APARTMENT NO.': apt_name,
                    'APARTMENT AREA (m²)': round(apartment_area_m2, 2),
                    'FULL TILES': full_count,
                    'IRREGULAR TILES': irregular_count,
                    'CUT X TILES': cut_x_count,
                    'CUT Y TILES': cut_y_count,
                    'ALL CUT TILES': all_cut_count,
                    'TOTAL CUT TILES': total_cuts,
                    'MATCHED < HALF': matched_less_half,
                    'MATCHED > HALF': matched_more_half,
                    'UNMATCHED < HALF': unmatched_less_half,
                    'UNMATCHED > HALF': unmatched_more_half,
                    'DESIGN EQUIVALENTS': round(design_total_equivalents, 1),
                    'OPTIMISED EQUIVALENTS': round(optimised_total_equivalents, 1),
                    'SMALL CUTS (<10mm)': small_cuts_count,
                    'TOTAL TILES': total_tiles,
                    'DESIGN TILING AREA (m²)': round(design_tiling_area_m2, 3),
                    'OPTIMISED TILING AREA (m²)': round(optimised_tiling_area_m2, 3),
                    'DESIGN WASTAGE %': round(design_wastage_percentage, 2),
                    'OPTIMISED WASTAGE %': round(optimised_wastage_percentage, 2),
                    'SAVINGS %': round(design_wastage_percentage - optimised_wastage_percentage, 2)
                })
            
            detailed_df = pd.DataFrame(detailed_breakdown)
            
            # Create tile specifications table
            tile_specs_data = [{
                'PARAMETER': 'TILE LENGTH (mm)',
                'VALUE': tile_length
            }, {
                'PARAMETER': 'TILE WIDTH (mm)', 
                'VALUE': tile_width
            }, {
                'PARAMETER': 'TILE AREA (mm²)',
                'VALUE': tile_area_mm2
            }, {
                'PARAMETER': 'TILE AREA (m²)',
                'VALUE': f"{tile_area_m2:.6f}"
            }, {
                'PARAMETER': 'HALF THRESHOLD (mm)',
                'VALUE': f"{half_threshold:.1f}"
            }]
            
            tile_specs_df = pd.DataFrame(tile_specs_data)
            
            # Create project summary with both design and optimised totals
            total_apartments = len(summary_data)
            total_apartment_area = sum(row['APARTMENT AREA (m²)'] for row in summary_data)
            total_tiles_all = sum(row['TOTAL TILES'] for row in detailed_breakdown)
            
            # Calculate project totals
            total_design_tiling_area = sum(row['DESIGN TILING AREA (m²)'] for row in summary_data)
            total_optimised_tiling_area = sum(row['OPTIMISED TILING AREA (m²)'] for row in summary_data)
            
            overall_design_wastage = ((total_design_tiling_area - total_apartment_area) / total_apartment_area) * 100 if total_apartment_area > 0 else 0
            overall_optimised_wastage = ((total_optimised_tiling_area - total_apartment_area) / total_apartment_area) * 100 if total_apartment_area > 0 else 0
            overall_savings = overall_design_wastage - overall_optimised_wastage
            
            project_summary_data = [{
                'METRIC': 'TOTAL APARTMENTS',
                'VALUE': total_apartments
            }, {
                'METRIC': 'TOTAL APARTMENT AREA (m²)',
                'VALUE': round(total_apartment_area, 2)
            }, {
                'METRIC': 'TOTAL TILES',
                'VALUE': total_tiles_all
            }, {
                'METRIC': 'DESIGN TILING AREA (m²)',
                'VALUE': round(total_design_tiling_area, 3)
            }, {
                'METRIC': 'OPTIMISED TILING AREA (m²)',
                'VALUE': round(total_optimised_tiling_area, 3)
            }, {
                'METRIC': 'DESIGN WASTAGE %',
                'VALUE': f"{round(overall_design_wastage, 2)}%"
            }, {
                'METRIC': 'OPTIMISED WASTAGE %',
                'VALUE': f"{round(overall_optimised_wastage, 2)}%"
            }, {
                'METRIC': 'TOTAL SAVINGS %',
                'VALUE': f"{round(overall_savings, 2)}%"
            }]
            
            project_summary_df = pd.DataFrame(project_summary_data)
            
            return {
                'summary_df': summary_df,
                'detailed_df': detailed_df,
                'tile_specs_df': tile_specs_df,
                'project_summary_df': project_summary_df
            }
            
        except Exception as e:
            print(f"Error creating enhanced summary data: {e}")
            return None